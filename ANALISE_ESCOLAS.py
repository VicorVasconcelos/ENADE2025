import pandas as pd
from openpyxl import load_workbook
import numpy as np
import os

 
# --- CONFIGURAÇÃO ---
# Caminhos
UPLOADS_PATH = r'C:\Users\victor.vasconcelos\OneDrive - CENTRO BRASILEIRO DE PESQUISA EM AVALIACAO E SELECAO E DE PROMOCAO DE EVENTOS - CEBRASPE\ENADE2025\Link Uploads Imagens Locais Indicados.xlsx'
SALAS_PATH = r'C:\Users\victor.vasconcelos\OneDrive - CENTRO BRASILEIRO DE PESQUISA EM AVALIACAO E SELECAO E DE PROMOCAO DE EVENTOS - CEBRASPE\ENADE2025\Relatório de salas.xlsx'
CONTROLE_PATH = r'C:\Users\victor.vasconcelos\OneDrive - CENTRO BRASILEIRO DE PESQUISA EM AVALIACAO E SELECAO E DE PROMOCAO DE EVENTOS - CEBRASPE\ENADE2025\ENAD - CONTROLE2.xlsx'
 
# UFs e Aba
UFS_RESPONSAVEIS = ['AC', 'AM', 'AP', 'PB', 'RR', 'RO']
ABA_RELATORIO = 'IndicacaoLocalProva'
 
 
# 1. LEITURA E PREPARAÇÃO DO RELATÓRIO DE LOCAIS
print("📊 Iniciando a leitura e preparação dos dados...")
try:
    df_locais = pd.read_excel(UPLOADS_PATH, sheet_name=ABA_RELATORIO, dtype=str)
    df_locais = df_locais.rename(columns={
        'IdLocalProva': 'Cod.Escola',
        'Cidade': 'Municipio',
        'LocalProva': 'Escola',
        'ResponsavelAlteracaoHomologacao': 'Responsavel'
    })
    df_locais = df_locais[df_locais['UF'].isin(UFS_RESPONSAVEIS)]
    df_locais = df_locais[df_locais['Responsavel'].isna()]
    df_locais = df_locais[['UF', 'Municipio', 'Cod.Escola', 'Escola']].drop_duplicates(subset='Cod.Escola')
except FileNotFoundError:
    print(f"❌ ERRO: Arquivo não encontrado em '{UPLOADS_PATH}'.")
    print("   Verifique se a pasta 'C:\\Planilhas' existe e se o arquivo está dentro dela.")
    exit()
except Exception as e:
    print(f"❌ ERRO ao ler o arquivo de locais: {e}")
    exit()
 
# 2. LEITURA E AGREGAÇÃO DO RELATÓRIO DE SALAS
print("🔄 Agregando informações das salas por escola...")
try:
    df_salas = pd.read_excel(SALAS_PATH, sheet_name=ABA_RELATORIO, dtype=str)
    df_salas = df_salas.rename(columns={'IdLocalProva': 'Cod.Escola'})
    df_salas['Capacidade'] = pd.to_numeric(df_salas['Capacidade'], errors='coerce').fillna(0)
    df_salas.dropna(subset=['Cod.Escola'], inplace=True)
   
    df_salas_agg = df_salas.groupby('Cod.Escola').agg(
        Qtd_Bloco=('Bloco', 'nunique'),
        Qtd_Sala=('Sala', 'count'),
        Capacidade_Total=('Capacidade', 'sum')
    ).reset_index()
 
    df_salas_primeira = df_salas.drop_duplicates(subset='Cod.Escola', keep='first').copy()
    df_salas_primeira['Metragem'] = df_salas_primeira.apply(
        lambda row: f"{row['Comprimento']} x {row['Largura']}" if pd.notna(row['Comprimento']) and pd.notna(row['Largura']) else 'N/A',
        axis=1
    )
    df_aptidao = df_salas.groupby('Cod.Escola')['AptoReceberAE'].apply(lambda x: 'Sim' if 'Sim' in x.values else 'Não').reset_index(name='Apta_AE')
    df_acessibilidade = df_salas.groupby('Cod.Escola')['PossuiAcessibilidade'].apply(lambda x: 'Sim' if 'Sim' in x.values else 'Não').reset_index(name='Acessibilidade')
 
    df_salas_final = pd.merge(df_salas_agg, df_salas_primeira[['Cod.Escola', 'Metragem']], on='Cod.Escola', how='left')
    df_salas_final = pd.merge(df_salas_final, df_aptidao, on='Cod.Escola', how='left')
    df_salas_final = pd.merge(df_salas_final, df_acessibilidade, on='Cod.Escola', how='left')
except FileNotFoundError:
    print(f"❌ ERRO: Arquivo não encontrado em '{SALAS_PATH}'.")
    print("   Verifique se a pasta 'C:\\Planilhas' existe e se o arquivo está dentro dela.")
    exit()
except Exception as e:
    print(f"❌ ERRO ao ler e agregar o arquivo de salas: {e}")
    exit()
 
# 3. UNIÃO DOS DADOS E ATUALIZAÇÃO DA PLANILHA DE CONTROLE
print("🔗 Unindo dados e preparando para atualização...")
df_final = pd.merge(df_locais, df_salas_final, on='Cod.Escola', how='left')
df_final.fillna({
    'Qtd_Bloco': 0, 'Qtd_Sala': 0, 'Capacidade_Total': 0,
    'Metragem': 'N/A', 'Apta_AE': 'Não', 'Acessibilidade': 'Não'
}, inplace=True)
 
try:
    wb = load_workbook(CONTROLE_PATH)
    contagem_ufs = {}
    print("✍️  Iniciando atualização da planilha de controle...")
 
    for uf in UFS_RESPONSAVEIS:
        if uf not in wb.sheetnames:
            print(f"⚠️  Aviso: Aba '{uf}' não encontrada. Pulando esta UF.")
            contagem_ufs[uf] = 0
            continue
 
        aba = wb[uf]
       
        try:
            df_existente = pd.read_excel(CONTROLE_PATH, sheet_name=uf, dtype=str, usecols="E", engine='openpyxl')
            df_existente.columns = ['Cod.Escola']
            codigos_existentes = set(df_existente['Cod.Escola'].dropna())
        except (ValueError, KeyError):
             codigos_existentes = set()
       
        df_novas = df_final[df_final['UF'] == uf]
        df_novas = df_novas[~df_novas['Cod.Escola'].isin(codigos_existentes)]
 
        if df_novas.empty:
            print(f"✅ Nenhuma escola nova para a UF {uf}.")
            contagem_ufs[uf] = 0
            continue
 
        proxima_linha = aba.max_row + 1
        for _, linha in df_novas.iterrows():
            # Colunas básicas (A-F)
            aba[f"A{proxima_linha}"] = "NOVO"
            aba[f"B{proxima_linha}"] = "EM ANÁLISE"
            aba[f"C{proxima_linha}"] = linha['UF']
            aba[f"D{proxima_linha}"] = linha['Municipio']
            aba[f"E{proxima_linha}"] = linha['Cod.Escola']
            aba[f"F{proxima_linha}"] = linha['Escola']
 
            # ==========================================================
            #      MAPEAMENTO DE COLUNAS FINAL - AJUSTADO
            # ==========================================================
            aba[f"I{proxima_linha}"] = int(linha['Qtd_Bloco'])
            aba[f"J{proxima_linha}"] = int(linha['Qtd_Sala'])
            aba[f"K{proxima_linha}"] = int(linha['Capacidade_Total'])
            aba[f"M{proxima_linha}"] = linha['Metragem']
            aba[f"O{proxima_linha}"] = linha['Apta_AE']
            aba[f"P{proxima_linha}"] = linha['Acessibilidade']
           
            proxima_linha += 1
 
        contagem_ufs[uf] = len(df_novas)
        print(f"👍 {len(df_novas)} nova(s) escola(s) adicionada(s) na aba '{uf}'.")
 
    wb.save(CONTROLE_PATH)
    print(f"\n💾 Planilha de controle '{CONTROLE_PATH}' salva com sucesso!")
 
except FileNotFoundError:
    print(f"❌ ERRO: Arquivo de controle não encontrado em: {CONTROLE_PATH}")
    print("   Verifique se a pasta 'C:\\Planilhas' existe e se o arquivo está dentro dela.")
    exit()
except Exception as e:
    print(f"❌ Ocorreu um erro inesperado ao manipular a planilha de controle: {e}")
 
# 4. RESUMO FINAL
print("\n--- RESUMO FINAL DA OPERAÇÃO ---")
total = sum(contagem_ufs.values())
for uf, qt in contagem_ufs.items():
    print(f"➡️  {uf}: {qt} escola(s) nova(s)")
print(f"\n🎯 Total geral: {total} escola(s) nova(s) adicionada(s).")
print("✅ Operação concluída com sucesso!")